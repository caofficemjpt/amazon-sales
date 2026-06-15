# Amazon Settlement Consolidator
# Merges Amazon Settlement TXT files + MTR CSV into a consolidated Excel report
# Dependencies: pip install pandas openpyxl
# Usage: python amazon_consolidator.py

"""
DATA ANALYSIS FINDINGS (pre-build):
-------------------------------------
Settlement TXT files (11 files, 95,442 total rows, 36 columns each):
  - All 11 files share the EXACT same 36-column header — no mismatches.
  - transaction-type unique values: 22 (Order, Refund, Amazon Easy Ship Charges,
    ServiceFee, Fulfillment Fee Refund, Storage Fee, RemovalComplete, etc.)
  - price-type unique values: 12 (Principal, Product Tax, Shipping, Shipping tax,
    TCS-IGST, TCS-CGST, TCS-SGST, TDS (Section 194-O), Gift Wrap Tax, Gift wrap)
  - item-related-fee-type: 26 values (FBA Pick & Pack Fee, FBA Weight Handling Fee,
    Fixed closing fee, Commission, Refund commission, Easy Ship weight handling fees, etc.)
  - promotion-type: 5 values (Promo rebates, Product tax discount, Shipping discount,
    Shipping tax discount)

  *** CRITICAL FINDING: Promo amounts are stored in the 'promotion-amount' column,
      NOT 'other-amount'. The 'other-amount' column is ALWAYS empty for promo rows.
      All 7,598 promo rows have their value in 'promotion-amount'. ***

MTR CSV (5,574 rows, 78 columns):
  - Transaction Type values: Shipment, Refund, FreeReplacement, Cancel
  - 1,079 Order Ids have more than 1 row (multi-item orders)
  - Multi-invoice rows share the same Invoice Number when it's a multi-item shipment
    (e.g., both items on the same invoice) — proportional split by Invoice Amount applies.

Cross-reference:
  - 5,595 unique order-ids in settlement; 4,248 unique Order Ids in MTR
  - 954 Order Ids are multi-invoice AND appear in settlement
  - Join key: MTR['Order Id'] == settlement['order-id']

Edge cases noted:
  - Promos: use 'promotion-amount' column (not 'other-amount' as spec states)
  - Some transaction-types (Storage Fee, RemovalComplete, etc.) appear in settlement
    but have no order-id — skip them in aggregation (blank order-id filter)
  - Refund rows in MTR have negative Invoice Amount — proportional split still works correctly
"""

from __future__ import annotations

import os
import threading
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
EXPECTED_SETTLEMENT_COLUMNS: list[str] = [
    "settlement-id", "settlement-start-date", "settlement-end-date", "deposit-date",
    "total-amount", "currency", "transaction-type", "order-id", "merchant-order-id",
    "adjustment-id", "shipment-id", "marketplace-name", "shipment-fee-type",
    "shipment-fee-amount", "order-fee-type", "order-fee-amount", "fulfillment-id",
    "posted-date", "order-item-code", "merchant-order-item-id",
    "merchant-adjustment-item-id", "sku", "quantity-purchased", "price-type",
    "price-amount", "item-related-fee-type", "item-related-fee-amount",
    "misc-fee-amount", "other-fee-amount", "other-fee-reason-description",
    "promotion-id", "promotion-type", "promotion-amount", "direct-payment-type",
    "direct-payment-amount", "other-amount",
]

REQUIRED_MTR_COLUMNS: list[str] = [
    "Seller Gstin", "Invoice Number", "Invoice Date", "Transaction Type",
    "Order Id", "Order Date", "Item Description", "Sku", "Quantity",
    "Tax Exclusive Gross", "Total Tax Amount", "Invoice Amount",
]

TCS_PRICE_TYPES: frozenset[str] = frozenset(
    ["TCS-IGST", "TCS-CGST", "TCS-SGST", "TCS-UTGST"]
)

TDS_PRICE_TYPE: str = "TDS (Section 194-O)"

# Output column order (used for validation and ordering)
OUTPUT_COLUMNS: list[str] = [
    "Seller GSTN", "Invoice No", "Invoice Date", "Transaction Type",
    "Order ID", "Order Date", "Item Description", "SKU", "Quantity",
    "Tax Exclusive Amount", "Total Tax Amount", "Invoice Amount",
    "Charges", "TCS-IGST", "TDS", "Promos", "Total",
    "Settlement ID", "Deposit Date",
]

# Columns that feed into the Total formula
TOTAL_FEED_COLUMNS: list[str] = [
    "Invoice Amount", "Charges", "TCS-IGST", "TDS", "Promos"
]

DECIMAL_PLACES: int = 2
MAX_COLUMN_WIDTH: int = 60

# Colour palette
HEADER_FILL_COLOR: str = "1F3864"    # dark navy
HEADER_FONT_COLOR: str = "FFFFFF"    # white
TOTAL_FILL_COLOR: str = "E8F0FE"     # light blue tint for Total column
SUMMARY_HDR_COLOR: str = "2E4D8A"    # slightly lighter navy for summary headers
SUMMARY_TOTAL_COLOR: str = "F2F7FF"  # very light blue for summary totals row

# ---------------------------------------------------------------------------
# Processing logic
# ---------------------------------------------------------------------------


def load_settlement_files(
    folder: str,
    log: callable,
    update_status: callable,
) -> pd.DataFrame:
    """Load and concatenate all TXT settlement files from *folder*."""
    txt_files = sorted(
        [f for f in os.listdir(folder) if f.lower().endswith(".txt")]
    )
    if not txt_files:
        raise FileNotFoundError(
            f"No .txt files found in folder: {folder}"
        )

    log(f"Found {len(txt_files)} settlement file(s).")
    frames: list[pd.DataFrame] = []

    for fname in txt_files:
        fpath = os.path.join(folder, fname)
        update_status(f"Loading {fname}…")
        try:
            df = pd.read_csv(
                fpath, sep="\t", dtype=str, encoding="utf-8-sig",
                encoding_errors="replace"
            )
        except Exception:
            df = pd.read_csv(
                fpath, sep="\t", dtype=str, encoding="utf-8",
                encoding_errors="replace"
            )

        # Strip BOM from column names if present
        df.columns = [c.strip().lstrip("\ufeff") for c in df.columns]
        log(f"  Loaded {fname}: {len(df):,} rows, {len(df.columns)} columns.")
        frames.append(df)

    settlement_df = pd.concat(frames, ignore_index=True)
    log(f"Total settlement rows: {len(settlement_df):,}")
    return settlement_df


def validate_settlement_headers(settlement_df: pd.DataFrame, log: callable) -> None:
    """Warn about any unexpected or missing columns in the settlement dataframe."""
    actual_set = set(settlement_df.columns)
    expected_set = set(EXPECTED_SETTLEMENT_COLUMNS)

    missing = expected_set - actual_set
    extra = actual_set - expected_set

    if missing:
        log(f"  WARNING — Settlement files missing expected columns: {sorted(missing)}")
    if extra:
        log(f"  NOTE — Settlement files contain extra columns: {sorted(extra)}")
    if not missing and not extra:
        log("  All settlement columns validated — headers match expected 36-column structure.")


def aggregate_settlement(
    settlement_df: pd.DataFrame, log: callable
) -> pd.DataFrame:
    """
    Aggregate settlement data per order-id into:
      Charges, TCS-IGST, TDS, Promos, Settlement ID, Deposit Date.

    NOTE: Per data analysis, promo amounts live in 'promotion-amount', not 'other-amount'.
    Deposit Date only appears on the first (header) row of each file (blank order-id).
    We capture it per settlement-id before filtering, then join it back.
    """
    # Convert numeric columns
    for col in ("price-amount", "item-related-fee-amount", "promotion-amount", "other-amount"):
        if col in settlement_df.columns:
            settlement_df[col] = pd.to_numeric(
                settlement_df[col], errors="coerce"
            ).fillna(0.0)

    # ── Extract deposit-date mapping BEFORE filtering blanks ──
    # deposit-date is only set on the very first row of each settlement file
    # (the row where order-id is blank and deposit-date is non-empty).
    sid_col = "settlement-id" if "settlement-id" in settlement_df.columns else None
    dd_col = "deposit-date" if "deposit-date" in settlement_df.columns else None
    settlement_id_to_deposit_date: dict[str, str] = {}
    if sid_col and dd_col:
        header_rows = settlement_df[
            settlement_df[dd_col].notna() & (settlement_df[dd_col].str.strip() != "")
        ]
        for _, row in header_rows.iterrows():
            sid = str(row[sid_col]).strip()
            dd = str(row[dd_col]).strip()
            if sid and dd:
                settlement_id_to_deposit_date[sid] = dd
        log(f"  Captured deposit dates for {len(settlement_id_to_deposit_date)} settlement file(s).")

    # Filter out rows with blank order-id
    valid = settlement_df["order-id"].notna() & (settlement_df["order-id"].str.strip() != "")
    df = settlement_df[valid].copy()
    log(
        f"  Settlement rows with valid order-id: {len(df):,} "
        f"(skipped {(~valid).sum():,} blank-order-id rows)"
    )

    # --- Charges: sum of item-related-fee-amount (all non-zero rows) ---
    charge_mask = df["item-related-fee-amount"] != 0.0
    charges = (
        df[charge_mask]
        .groupby("order-id", sort=False)["item-related-fee-amount"]
        .sum()
        .rename("Charges")
    )

    # --- TCS-IGST: sum of price-amount where price-type in TCS set ---
    tcs_mask = df["price-type"].isin(TCS_PRICE_TYPES)
    tcs = (
        df[tcs_mask]
        .groupby("order-id", sort=False)["price-amount"]
        .sum()
        .rename("TCS-IGST")
    )

    # --- TDS: sum of price-amount where price-type == TDS_PRICE_TYPE ---
    tds_mask = df["price-type"] == TDS_PRICE_TYPE
    tds = (
        df[tds_mask]
        .groupby("order-id", sort=False)["price-amount"]
        .sum()
        .rename("TDS")
    )

    # --- Promos: sum of promotion-amount where promotion-type is non-empty ---
    promo_mask = (
        df["promotion-type"].notna() & (df["promotion-type"].str.strip() != "")
    )
    promo_col = "promotion-amount" if "promotion-amount" in df.columns else "other-amount"
    promos = (
        df[promo_mask]
        .groupby("order-id", sort=False)[promo_col]
        .sum()
        .rename("Promos")
    )

    # --- Settlement ID: first value per order-id ---
    if sid_col:
        settlement_ids = (
            df.groupby("order-id", sort=False)[sid_col]
            .first()
            .rename("Settlement ID")
        )
    else:
        settlement_ids = pd.Series(dtype=str, name="Settlement ID")

    summary = (
        pd.DataFrame({
            "Charges": charges,
            "TCS-IGST": tcs,
            "TDS": tds,
            "Promos": promos,
            "Settlement ID": settlement_ids,
        })
        .reset_index()
        .rename(columns={"order-id": "order-id"})
    )

    # Fill numeric cols with 0
    for col in ("Charges", "TCS-IGST", "TDS", "Promos"):
        summary[col] = summary[col].fillna(0.0)
    summary["Settlement ID"] = summary["Settlement ID"].fillna("").astype(str)

    # --- Deposit Date: map from settlement-id → deposit-date ---
    summary["Deposit Date"] = (
        summary["Settlement ID"]
        .str.strip()
        .map(settlement_id_to_deposit_date)
        .fillna("")
    )

    log(f"  Aggregated settlement: {len(summary):,} unique order-ids")
    return summary


def load_mtr(mtr_path: str, log: callable, update_status: callable) -> pd.DataFrame:
    """Load the MTR CSV and validate required columns."""
    update_status("Loading MTR CSV…")
    try:
        mtr = pd.read_csv(mtr_path, dtype=str, encoding="utf-8",
                          encoding_errors="replace")
    except Exception:
        mtr = pd.read_csv(mtr_path, dtype=str, encoding="utf-8-sig",
                          encoding_errors="replace")

    mtr.columns = [c.strip() for c in mtr.columns]
    log(f"  MTR loaded: {len(mtr):,} rows, {len(mtr.columns)} columns.")

    missing_cols = [c for c in REQUIRED_MTR_COLUMNS if c not in mtr.columns]
    if missing_cols:
        raise ValueError(
            f"MTR CSV is missing required columns: {missing_cols}"
        )

    # Parse numeric MTR columns
    for col in ("Invoice Amount", "Tax Exclusive Gross", "Total Tax Amount"):
        mtr[col] = pd.to_numeric(mtr[col], errors="coerce").fillna(0.0)
    mtr["Quantity"] = pd.to_numeric(mtr["Quantity"], errors="coerce").fillna(0).astype(int)

    return mtr


def join_and_split(
    mtr: pd.DataFrame,
    settlement_summary: pd.DataFrame,
    log: callable,
    update_status: callable,
) -> tuple[pd.DataFrame, int, int]:
    """
    Left-join MTR → settlement summary, then split settlement values proportionally
    for multi-invoice orders.
    """
    update_status("Joining datasets…")

    sett = settlement_summary.rename(columns={"order-id": "_order_id_key"})
    merged = mtr.merge(
        sett,
        how="left",
        left_on="Order Id",
        right_on="_order_id_key",
    )
    merged.drop(columns=["_order_id_key"], inplace=True)

    # Fill NaN settlement numeric columns with 0
    for col in ("Charges", "TCS-IGST", "TDS", "Promos"):
        merged[col] = pd.to_numeric(merged[col], errors="coerce").fillna(0.0)

    # Fill string settlement columns with ""
    for col in ("Settlement ID", "Deposit Date"):
        if col in merged.columns:
            merged[col] = merged[col].fillna("")
        else:
            merged[col] = ""

    # Count unmatched orders
    settlement_order_ids = set(settlement_summary["order-id"].dropna())
    mtr_order_ids = set(mtr["Order Id"].dropna())
    no_match_count = len(mtr_order_ids - settlement_order_ids - {""})
    settlement_no_mtr = len(settlement_order_ids - mtr_order_ids - {""})

    log(f"  MTR orders with no settlement match: {no_match_count:,}")
    log(f"  Settlement orders with no MTR match: {settlement_no_mtr:,} (excluded from output)")

    # --- Proportional split for multi-invoice orders ---
    update_status("Splitting charges for multi-invoice orders…")

    order_invoice_totals = (
        mtr.groupby("Order Id", sort=False)["Invoice Amount"]
        .sum()
        .rename("_order_invoice_total")
    )
    merged = merged.merge(order_invoice_totals, on="Order Id", how="left")

    order_counts = mtr["Order Id"].value_counts()
    multi_invoice_orders = set(order_counts[order_counts > 1].index)
    multi_count = len(multi_invoice_orders)
    log(f"  Multi-invoice orders requiring proportional split: {multi_count:,}")

    multi_mask = merged["Order Id"].isin(multi_invoice_orders)
    order_total = merged["_order_invoice_total"]
    row_amount = merged["Invoice Amount"]

    weight = pd.Series(0.0, index=merged.index)
    safe_mask = multi_mask & (order_total != 0) & order_total.notna()
    weight[safe_mask] = row_amount[safe_mask] / order_total[safe_mask]
    weight[~multi_mask] = 1.0

    for col in ("Charges", "TCS-IGST", "TDS", "Promos"):
        merged[col] = (merged[col] * weight).round(DECIMAL_PLACES)

    merged.drop(columns=["_order_invoice_total"], inplace=True)

    return merged, no_match_count, settlement_no_mtr


def build_output(merged: pd.DataFrame) -> pd.DataFrame:
    """Select, rename and order columns for the final output DataFrame."""
    output = pd.DataFrame()
    output["Seller GSTN"] = merged["Seller Gstin"]
    output["Invoice No"] = merged["Invoice Number"]
    output["Invoice Date"] = merged["Invoice Date"]
    output["Transaction Type"] = merged["Transaction Type"]
    output["Order ID"] = merged["Order Id"]
    output["Order Date"] = merged["Order Date"]
    output["Item Description"] = merged["Item Description"]
    output["SKU"] = merged["Sku"]
    output["Quantity"] = merged["Quantity"].astype(int)
    output["Tax Exclusive Amount"] = merged["Tax Exclusive Gross"]
    output["Total Tax Amount"] = merged["Total Tax Amount"]
    output["Invoice Amount"] = merged["Invoice Amount"]
    output["Charges"] = merged["Charges"].fillna(0.0)
    output["TCS-IGST"] = merged["TCS-IGST"].fillna(0.0)
    output["TDS"] = merged["TDS"].fillna(0.0)
    output["Promos"] = merged["Promos"].fillna(0.0)

    # Total = Invoice Amount + Charges + TCS-IGST + TDS + Promos
    output["Total"] = (
        output["Invoice Amount"]
        + output["Charges"]
        + output["TCS-IGST"]
        + output["TDS"]
        + output["Promos"]
    ).round(DECIMAL_PLACES)

    output["Settlement ID"] = merged.get("Settlement ID", pd.Series("", index=merged.index)).fillna("")
    output["Deposit Date"] = merged.get("Deposit Date", pd.Series("", index=merged.index)).fillna("")

    return output


def build_summary(output_df: pd.DataFrame, month: str, year: str) -> pd.DataFrame:
    """
    Build a GSTN-wise summary DataFrame for the given period.

    Columns per GSTN:
      - Taxable        = sum of Tax Exclusive Amount
      - GST            = sum of Total Tax Amount
      - Total          = sum of Invoice Amount
      - Charges        = sum of Charges
      - Total Received = sum of Total (Invoice Amount + Charges + TCS + TDS + Promos)
    """
    grp = output_df.groupby("Seller GSTN", sort=True)
    summary = pd.DataFrame({
        "Seller GSTN": grp["Seller GSTN"].first(),
        "Taxable": grp["Tax Exclusive Amount"].sum().round(DECIMAL_PLACES),
        "GST": grp["Total Tax Amount"].sum().round(DECIMAL_PLACES),
        "Total": grp["Invoice Amount"].sum().round(DECIMAL_PLACES),
        "Charges": grp["Charges"].sum().round(DECIMAL_PLACES),
        "Total Received Amount": grp["Total"].sum().round(DECIMAL_PLACES),
    }).reset_index(drop=True)

    # Append grand-total row
    total_row = pd.DataFrame([{
        "Seller GSTN": "GRAND TOTAL",
        "Taxable": summary["Taxable"].sum().round(DECIMAL_PLACES),
        "GST": summary["GST"].sum().round(DECIMAL_PLACES),
        "Total": summary["Total"].sum().round(DECIMAL_PLACES),
        "Charges": summary["Charges"].sum().round(DECIMAL_PLACES),
        "Total Received Amount": summary["Total Received Amount"].sum().round(DECIMAL_PLACES),
    }])
    summary = pd.concat([summary, total_row], ignore_index=True)
    return summary


def _apply_sheet_styles(
    ws,
    df: pd.DataFrame,
    amount_col_names: list[str],
    qty_col_name: str | None = None,
    highlight_last_row: bool = False,
    highlight_col_name: str | None = None,
) -> None:
    """
    Apply common Excel styles to a worksheet:
      - Styled header row (navy background, white bold text)
      - Freeze top row
      - Number formats for amounts and quantity
      - Auto-fit column widths
      - Optional highlight on a specific column (e.g., Total)
      - Optional bold+shaded last row (for grand-total)
    """
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL_COLOR)
    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Build column index lookup
    col_name_to_idx: dict[str, int] = {
        cell.value: cell.column for cell in ws[1]
    }

    # Header row style
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    ws.freeze_panes = "A2"

    # Auto-fit column widths
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            df[col_name].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[col_letter].width = min(int(max_len) + 2, MAX_COLUMN_WIDTH)

    # Amount column number format
    for col_name in amount_col_names:
        if col_name in col_name_to_idx:
            col_letter = get_column_letter(col_name_to_idx[col_name])
            for cell in ws[col_letter]:
                if cell.row > 1:
                    cell.number_format = "#,##0.00"

    # Quantity column format
    if qty_col_name and qty_col_name in col_name_to_idx:
        col_letter = get_column_letter(col_name_to_idx[qty_col_name])
        for cell in ws[col_letter]:
            if cell.row > 1:
                cell.number_format = "0"

    # Highlight the Total column with a light blue tint
    if highlight_col_name and highlight_col_name in col_name_to_idx:
        total_fill = PatternFill(fill_type="solid", fgColor=TOTAL_FILL_COLOR)
        col_letter = get_column_letter(col_name_to_idx[highlight_col_name])
        for cell in ws[col_letter]:
            if cell.row > 1:
                cell.fill = total_fill

    # Bold + shaded last row (grand total)
    if highlight_last_row and ws.max_row > 1:
        grand_fill = PatternFill(fill_type="solid", fgColor="D6E4F7")
        grand_font = Font(bold=True, size=10)
        for cell in ws[ws.max_row]:
            cell.fill = grand_fill
            cell.font = grand_font
            cell.alignment = Alignment(horizontal="center")


def write_excel(
    output_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    output_path: str,
    month: str,
    year: str,
    existing_path: str,
    log: callable,
    update_status: callable,
) -> None:
    """
    Write the consolidated and summary sheets to an Excel file.

    If *existing_path* is provided and the file exists, load it and append
    new rows to the 'Consolidated' sheet (avoiding duplicate Invoice No rows),
    then regenerate the summary sheet.
    """
    update_status("Writing Excel output…")

    AMOUNT_COLS = [
        "Tax Exclusive Amount", "Total Tax Amount", "Invoice Amount",
        "Charges", "TCS-IGST", "TDS", "Promos", "Total",
    ]
    SUMMARY_AMOUNT_COLS = ["Taxable", "GST", "Total", "Charges", "Total Received Amount"]
    summary_sheet_name = f"Summary of {month} - {year}"

    # ── Determine if we append to an existing file ──
    if existing_path and os.path.isfile(existing_path):
        log(f"  Existing file detected — appending new rows to: {existing_path}")
        try:
            existing_consolidated = pd.read_excel(
                existing_path, sheet_name="Consolidated", dtype=str
            )
            # Coerce numeric cols in existing data
            for col in AMOUNT_COLS:
                if col in existing_consolidated.columns:
                    existing_consolidated[col] = pd.to_numeric(
                        existing_consolidated[col], errors="coerce"
                    ).fillna(0.0)
            if "Quantity" in existing_consolidated.columns:
                existing_consolidated["Quantity"] = (
                    pd.to_numeric(existing_consolidated["Quantity"], errors="coerce")
                    .fillna(0)
                    .astype(int)
                )

            # Deduplicate: drop rows from existing that also appear in new data
            existing_invoice_nos = set(existing_consolidated["Invoice No"].dropna())
            new_invoice_nos = set(output_df["Invoice No"].dropna())
            overlap = existing_invoice_nos & new_invoice_nos
            if overlap:
                log(f"  WARNING — {len(overlap):,} Invoice No(s) already in existing file; they will be overwritten with new data.")
                existing_consolidated = existing_consolidated[
                    ~existing_consolidated["Invoice No"].isin(overlap)
                ]

            # Align columns: ensure both frames have the same columns in OUTPUT_COLUMNS order
            for col in OUTPUT_COLUMNS:
                if col not in existing_consolidated.columns:
                    existing_consolidated[col] = "" if col in ("Settlement ID", "Deposit Date") else 0.0
                if col not in output_df.columns:
                    output_df[col] = "" if col in ("Settlement ID", "Deposit Date") else 0.0

            combined_df = pd.concat(
                [existing_consolidated[OUTPUT_COLUMNS], output_df[OUTPUT_COLUMNS]],
                ignore_index=True,
            )
            log(f"  Combined: {len(existing_consolidated):,} existing + {len(output_df):,} new = {len(combined_df):,} total rows")

            # Rebuild summary over ALL data (not just new)
            # Re-coerce numeric so summary is accurate
            for col in AMOUNT_COLS:
                if col in combined_df.columns:
                    combined_df[col] = pd.to_numeric(combined_df[col], errors="coerce").fillna(0.0)
            combined_df["Quantity"] = (
                pd.to_numeric(combined_df["Quantity"], errors="coerce").fillna(0).astype(int)
            )
            combined_summary = build_summary(combined_df, month, year)

            final_consolidated = combined_df
            final_summary = combined_summary
        except Exception as exc:
            log(f"  WARNING — Could not read existing file ({exc}). Writing fresh output.")
            final_consolidated = output_df[OUTPUT_COLUMNS].copy()
            final_summary = summary_df
    else:
        final_consolidated = output_df[OUTPUT_COLUMNS].copy()
        final_summary = summary_df

    # ── Write the Excel file ──
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        final_consolidated.to_excel(writer, sheet_name="Consolidated", index=False)
        final_summary.to_excel(writer, sheet_name=summary_sheet_name, index=False)

        # Style Consolidated sheet
        ws_cons = writer.sheets["Consolidated"]
        _apply_sheet_styles(
            ws_cons,
            final_consolidated,
            amount_col_names=AMOUNT_COLS,
            qty_col_name="Quantity",
            highlight_last_row=False,
            highlight_col_name="Total",
        )

        # Style Summary sheet
        ws_summ = writer.sheets[summary_sheet_name]
        _apply_sheet_styles(
            ws_summ,
            final_summary,
            amount_col_names=SUMMARY_AMOUNT_COLS,
            qty_col_name=None,
            highlight_last_row=True,
            highlight_col_name="Total Received Amount",
        )

    log(f"  Excel file written: {output_path}")
    log(f"  Sheets: 'Consolidated' ({len(final_consolidated):,} rows), '{summary_sheet_name}' ({len(final_summary) - 1:,} GSTN rows + grand total)")


def run_processing(
    settlement_folder: str,
    mtr_path: str,
    output_path: str,
    existing_path: str,
    month: str,
    year: str,
    log: callable,
    update_status: callable,
    on_complete: callable,
    on_error: callable,
) -> None:
    """Main processing pipeline — runs in a background thread."""
    try:
        # Step 1: Load settlement files
        update_status("Loading settlement files…")
        log("━━━ Step 1: Loading settlement files… ━━━")
        settlement_df = load_settlement_files(settlement_folder, log, update_status)

        # Step 2: Validate headers
        update_status("Validating headers…")
        log("━━━ Step 2: Validating headers… ━━━")
        validate_settlement_headers(settlement_df, log)

        # Step 3: Aggregate settlement data
        update_status("Aggregating settlement data by Order ID…")
        log("━━━ Step 3: Aggregating settlement data by Order ID… ━━━")
        settlement_summary = aggregate_settlement(settlement_df, log)

        # Step 4: Load MTR CSV
        log("━━━ Step 4: Loading MTR CSV… ━━━")
        mtr = load_mtr(mtr_path, log, update_status)

        # Step 5 + 6: Join and proportional split
        log("━━━ Step 5: Joining datasets… ━━━")
        merged, no_match_count, settlement_no_mtr = join_and_split(
            mtr, settlement_summary, log, update_status
        )
        log("━━━ Step 6: Proportional splitting complete ━━━")

        # Step 7: Build output columns
        log("━━━ Step 7: Building output structure… ━━━")
        update_status("Building output columns…")
        output_df = build_output(merged)

        # Step 7b: Build summary
        log("━━━ Step 7b: Building GSTN-wise summary… ━━━")
        summary_df = build_summary(output_df, month, year)
        log(f"  Summary covers {len(summary_df) - 1} GSTN(s).")

        # Step 8: Write Excel
        log("━━━ Step 8: Writing Excel output… ━━━")
        write_excel(
            output_df, summary_df, output_path,
            month, year, existing_path,
            log, update_status,
        )

        total_rows = len(output_df)
        summary_msg = (
            f"✅ Done! {total_rows:,} new rows processed. "
            f"{no_match_count:,} MTR orders had no settlement match. "
            f"{settlement_no_mtr:,} settlement orders had no MTR match."
        )
        log(summary_msg)
        on_complete(summary_msg)

    except Exception as exc:
        tb = traceback.format_exc()
        log(f"\n❌ ERROR: {exc}\n\nTraceback:\n{tb}")
        on_error(str(exc))


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

MONTHS: list[str] = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


class ConsolidatorApp(tk.Tk):
    """Main application window for the Amazon Settlement Consolidator."""

    def __init__(self) -> None:
        super().__init__()
        self.title("Amazon Settlement Consolidator")
        self.geometry("820x720")
        self.minsize(740, 620)
        self.configure(bg="#F0F4F8")
        self._build_ui()
        self._center_window()

    def _center_window(self) -> None:
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    # ------------------------------------------------------------------
    # UI Construction
    # ------------------------------------------------------------------

    def _build_ui(self) -> None:
        # ── Title bar ──
        title_frame = tk.Frame(self, bg="#1F3864", pady=12)
        title_frame.pack(fill="x")
        tk.Label(
            title_frame,
            text="🛒  Amazon Settlement Consolidator",
            bg="#1F3864", fg="white",
            font=("Segoe UI", 16, "bold"),
        ).pack()
        tk.Label(
            title_frame,
            text="Merge Settlement TXT reports + MTR CSV into a single Excel file",
            bg="#1F3864", fg="#A8C0D6",
            font=("Segoe UI", 9),
        ).pack()

        # ── Input files frame ──
        input_frame = tk.LabelFrame(
            self, text="Input Files", bg="#F0F4F8", fg="#1F3864",
            font=("Segoe UI", 10, "bold"), padx=16, pady=10,
        )
        input_frame.pack(fill="x", padx=20, pady=(14, 6))
        input_frame.columnconfigure(1, weight=1)

        self._settlement_var = tk.StringVar()
        self._add_file_row(input_frame, 0, "Settlement Folder:", self._settlement_var,
                           "📁 Browse", self._pick_settlement_folder)

        self._mtr_var = tk.StringVar()
        self._add_file_row(input_frame, 1, "MTR CSV File:", self._mtr_var,
                           "📁 Browse", self._pick_mtr_file)

        self._existing_var = tk.StringVar()
        self._add_file_row(input_frame, 2, "Existing File (opt):", self._existing_var,
                           "📁 Browse", self._pick_existing_file,
                           hint="Optional — new rows will be appended here")

        self._output_var = tk.StringVar(value="consolidated_output.xlsx")
        self._add_file_row(input_frame, 3, "Output Excel File:", self._output_var,
                           "💾 Save As", self._pick_output_file)

        # ── Period frame ──
        period_frame = tk.LabelFrame(
            self, text="Report Period", bg="#F0F4F8", fg="#1F3864",
            font=("Segoe UI", 10, "bold"), padx=16, pady=10,
        )
        period_frame.pack(fill="x", padx=20, pady=(0, 6))

        # Month dropdown
        tk.Label(period_frame, text="Month:", bg="#F0F4F8", fg="#333333",
                 font=("Segoe UI", 9)).grid(row=0, column=0, sticky="e", padx=(0, 8), pady=4)

        self._month_var = tk.StringVar(value="April")
        month_menu = tk.OptionMenu(period_frame, self._month_var, *MONTHS)
        month_menu.config(font=("Segoe UI", 9), bg="white", relief="solid",
                          bd=1, highlightthickness=0, cursor="hand2", width=14)
        month_menu["menu"].config(font=("Segoe UI", 9))
        month_menu.grid(row=0, column=1, sticky="w", pady=4)

        # Year entry
        tk.Label(period_frame, text="Year:", bg="#F0F4F8", fg="#333333",
                 font=("Segoe UI", 9)).grid(row=0, column=2, sticky="e",
                                            padx=(24, 8), pady=4)

        import datetime
        self._year_var = tk.StringVar(value=str(datetime.date.today().year))
        year_entry = tk.Entry(period_frame, textvariable=self._year_var,
                              font=("Segoe UI", 9), width=8,
                              relief="solid", bd=1, bg="white")
        year_entry.grid(row=0, column=3, sticky="w", pady=4)

        # ── Process button ──
        btn_frame = tk.Frame(self, bg="#F0F4F8")
        btn_frame.pack(pady=8)
        self._process_btn = tk.Button(
            btn_frame,
            text="▶  Process Files",
            command=self._start_processing,
            bg="#1F3864", fg="white",
            font=("Segoe UI", 11, "bold"),
            relief="flat", padx=28, pady=8,
            cursor="hand2",
            activebackground="#2E4D8A", activeforeground="white",
        )
        self._process_btn.pack()
        self._process_btn.bind("<Enter>", lambda _: self._process_btn.config(bg="#2E4D8A"))
        self._process_btn.bind("<Leave>", lambda _: self._process_btn.config(bg="#1F3864"))

        # ── Log window ──
        log_frame = tk.LabelFrame(
            self, text="Log / Status", bg="#F0F4F8", fg="#1F3864",
            font=("Segoe UI", 10, "bold"), padx=8, pady=8,
        )
        log_frame.pack(fill="both", expand=True, padx=20, pady=(4, 4))

        self._log_text = scrolledtext.ScrolledText(
            log_frame, wrap="word", state="disabled",
            font=("Consolas", 9), bg="#FAFAFA", fg="#1A1A2E",
            relief="flat", bd=1,
        )
        self._log_text.pack(fill="both", expand=True)

        # ── Status bar ──
        self._status_var = tk.StringVar(value="Status: Ready")
        tk.Label(
            self, textvariable=self._status_var,
            bg="#DDE6F0", fg="#1F3864",
            font=("Segoe UI", 9), anchor="w", padx=12, pady=4,
        ).pack(fill="x", side="bottom")

    def _add_file_row(
        self, parent: tk.Widget, row: int, label: str,
        var: tk.StringVar, btn_text: str, picker: callable,
        hint: str = "",
    ) -> None:
        tk.Label(
            parent, text=label, bg="#F0F4F8", fg="#333333",
            font=("Segoe UI", 9), anchor="e", width=20,
        ).grid(row=row, column=0, sticky="e", pady=5, padx=(0, 8))

        entry_frame = tk.Frame(parent, bg="#F0F4F8")
        entry_frame.grid(row=row, column=1, sticky="ew", pady=5)
        entry_frame.columnconfigure(0, weight=1)

        tk.Entry(
            entry_frame, textvariable=var, font=("Segoe UI", 9),
            relief="solid", bd=1, bg="white",
        ).grid(row=0, column=0, sticky="ew")

        if hint:
            tk.Label(
                entry_frame, text=hint, bg="#F0F4F8", fg="#888888",
                font=("Segoe UI", 7, "italic"), anchor="w",
            ).grid(row=1, column=0, sticky="w")

        tk.Button(
            parent, text=btn_text, command=picker,
            bg="#E8EEF4", fg="#1F3864",
            font=("Segoe UI", 9), relief="flat", padx=8, pady=4,
            cursor="hand2",
        ).grid(row=row, column=2, sticky="w", pady=5, padx=(8, 0))

    # ------------------------------------------------------------------
    # Pickers
    # ------------------------------------------------------------------

    def _pick_settlement_folder(self) -> None:
        folder = filedialog.askdirectory(title="Select Settlement Folder")
        if folder:
            self._settlement_var.set(folder)

    def _pick_mtr_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Select MTR CSV File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if path:
            self._mtr_var.set(path)

    def _pick_existing_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Select Existing Consolidated Excel File (optional)",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self._existing_var.set(path)

    def _pick_output_file(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Save Output Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="consolidated_output.xlsx",
        )
        if path:
            self._output_var.set(path)

    # ------------------------------------------------------------------
    # Logging helpers
    # ------------------------------------------------------------------

    def _log(self, message: str) -> None:
        """Append a line to the log widget (thread-safe via after())."""
        self.after(0, self._append_log, message)

    def _append_log(self, message: str) -> None:
        self._log_text.config(state="normal")
        self._log_text.insert("end", message + "\n")
        self._log_text.see("end")
        self._log_text.config(state="disabled")

    def _update_status(self, text: str) -> None:
        self.after(0, self._status_var.set, f"Status: {text}")

    # ------------------------------------------------------------------
    # Validation & Processing
    # ------------------------------------------------------------------

    def _validate_inputs(self) -> bool:
        settlement_folder = self._settlement_var.get().strip()
        mtr_path = self._mtr_var.get().strip()
        output_path = self._output_var.get().strip()
        year = self._year_var.get().strip()

        if not settlement_folder:
            messagebox.showerror("Missing Input", "Please select a Settlement Folder.")
            return False
        if not os.path.isdir(settlement_folder):
            messagebox.showerror("Invalid Folder", f"Settlement folder not found:\n{settlement_folder}")
            return False
        txt_count = sum(1 for f in os.listdir(settlement_folder) if f.lower().endswith(".txt"))
        if txt_count == 0:
            messagebox.showerror(
                "No TXT Files",
                f"No .txt files found in:\n{settlement_folder}\n\n"
                "Please select a folder containing Amazon Settlement Report files.",
            )
            return False

        if not mtr_path:
            messagebox.showerror("Missing Input", "Please select an MTR CSV file.")
            return False
        if not os.path.isfile(mtr_path):
            messagebox.showerror("File Not Found", f"MTR CSV file not found:\n{mtr_path}")
            return False

        if not output_path:
            messagebox.showerror("Missing Input", "Please specify an output Excel file path.")
            return False
        output_dir = os.path.dirname(output_path) or "."
        if not os.path.isdir(output_dir):
            messagebox.showerror("Invalid Path", f"Output directory does not exist:\n{output_dir}")
            return False

        if not year.isdigit() or len(year) != 4:
            messagebox.showerror("Invalid Year", f"Year must be a 4-digit number (e.g., 2026). Got: {year!r}")
            return False

        return True

    def _start_processing(self) -> None:
        if not self._validate_inputs():
            return

        # Clear log
        self._log_text.config(state="normal")
        self._log_text.delete("1.0", "end")
        self._log_text.config(state="disabled")

        settlement_folder = self._settlement_var.get().strip()
        mtr_path = self._mtr_var.get().strip()
        output_path = self._output_var.get().strip()
        existing_path = self._existing_var.get().strip()
        month = self._month_var.get()
        year = self._year_var.get().strip()

        self._process_btn.config(state="disabled", text="⏳  Processing…")
        self._update_status("Processing — please wait…")
        self._log("Amazon Settlement Consolidator — starting…")
        self._log(f"Settlement folder : {settlement_folder}")
        self._log(f"MTR CSV           : {mtr_path}")
        self._log(f"Existing file     : {existing_path or '(none)'}")
        self._log(f"Output Excel      : {output_path}")
        self._log(f"Report period     : {month} {year}")
        self._log("")

        thread = threading.Thread(
            target=run_processing,
            kwargs={
                "settlement_folder": settlement_folder,
                "mtr_path": mtr_path,
                "output_path": output_path,
                "existing_path": existing_path,
                "month": month,
                "year": year,
                "log": self._log,
                "update_status": self._update_status,
                "on_complete": self._on_complete,
                "on_error": self._on_error,
            },
            daemon=True,
        )
        thread.start()

    def _on_complete(self, summary: str) -> None:
        self.after(0, self._handle_complete, summary)

    def _handle_complete(self, summary: str) -> None:
        self._process_btn.config(state="normal", text="▶  Process Files")
        self._update_status("Done!")
        messagebox.showinfo("Processing Complete", summary)

    def _on_error(self, error_msg: str) -> None:
        self.after(0, self._handle_error, error_msg)

    def _handle_error(self, error_msg: str) -> None:
        self._process_btn.config(state="normal", text="▶  Process Files")
        self._update_status("Error — check log for details")
        messagebox.showerror(
            "Processing Error",
            f"An error occurred:\n\n{error_msg}\n\nSee the log window for the full traceback.",
        )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app = ConsolidatorApp()
    app.mainloop()
